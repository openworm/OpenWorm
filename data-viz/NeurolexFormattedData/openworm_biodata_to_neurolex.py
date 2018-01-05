# -*- coding: utf-8 -*-
# <nbformat>3.0</nbformat>


#This file will help to align the data from an excel workbook into the proper format for uploading to Neurolex

import openpyxl, csv

CUfname = 'NeuroLex Cell Upload Template.xlsx'                      #from here: http://neurolex.org/wiki/CSVImportInstructions
biodatfname = 'neuron_transmitter_peptide_data_July2013.xlsx'       #from here: https://docs.google.com/spreadsheet/ccc?key=0Avt3mQaA-HaMdHZWLWJEMkdrSGtIX0FMUlh2LXhXWnc#gid=0

outfilename = 'NLX_formatted.csv'


def dictfromsheet(sheet):
    #sheet is an openpyxl sheet
    #from here: http://quuux.org/code/xlsx-dict-reader.py/
    rows = sheet.get_highest_row()
    cols = sheet.get_highest_column()

    headers = dict( (i, sheet.cell(row=0, column=i).value) for i in range(cols) )
    
    def item(i, j):
        return (sheet.cell(row=0, column=j).value,
                sheet.cell(row=i, column=j).value)
 
    return (dict(item(i,j) for j in range(cols)) for i in range(1, rows)), headers
 

#open cell upload template and grab headers for proper neurolex column names
with open(CUfname, 'rb') as f:
    tempbk = openpyxl.load_workbook(f)
    tempsheet = tempbk.get_sheet_by_name(name = 'Sheet1')

CUtempdat = []
CUdat, CUtemplate_headers = dictfromsheet(tempsheet)
for rown, row in enumerate(CUdat):
    CUtempdat.append(row)

#open the openworm biodata workbook
with open(biodatfname, 'rb') as f:
    biodatbk = openpyxl.load_workbook(f)
    biodat_sheetnames = biodatbk.get_sheet_names()


#go through sheet by sheet and assign columns to appropriate neurolex columns, reformatting as needed
sheetnum = 0 #for now explicitly munging in the 'Neurons' worksheet
Neurondat = []
currdat, currheaders = dictfromsheet(biodatbk.get_sheet_by_name(name = biodat_sheetnames[sheetnum]))
for rown, row in enumerate(currdat):
    Neurondat.append(row)

#Here's where you get your columns all pretty for Neurolex
NLX_data = [];
NLX_headers = CUtemplate_headers.values()
#NLX_headers.append(u'nametoappend')
for row in Neurondat:
    row_out = row
    
    ###rename misnamed columns  eg: row_out['newname'] = row_out.pop('oldname')
    #try to match these to the columns in the cell upload template from neurolex
    row_out['Label'] = row_out.pop('Neuron')
    row_out['Has role'] = row_out.pop('Type')
    row_out['Located in'] = row_out.pop('Location')
    row_out['Definition'] = row_out.pop('Description')
    row_out['Neurotransmitter'] = row_out.pop('Neurotransmitter/Neuropeptide')
    
    #for now, stick innexin expression and receptor expression both into MolecularConstituents
    ie = row_out.pop('Innexin Expression');
    if ie is None: ie = ''
    re = row_out.pop('Receptor expression');
    if re is None: re = ''
    row_out['MolecularConstituents'] = 'Innexin Expression: ' + ie + '   Receptor Expression: ' + re
    
    fe = row_out.pop('Function');
    if fe is None: fe = ''
    row_out['Comment'] = 'Function: ' + fe
    
    
    #add manual data - if adding a column, also add to NLX_headers above
    row_out['SuperCategory'] = 'OpenWorm 2013 Parcel'
    row_out['Species'] = 'Caenorhabditis_elegans'
    
    #delete unwanted columns
    #del row_out['nameofunwantedcolumn']
    
    NLX_data.append(row_out)
#done with 'Neurons' worksheet

with open(outfilename, "wb") as ofile:
    cw = csv.DictWriter(ofile, fieldnames=NLX_headers, delimiter=',', quoting=csv.QUOTE_ALL) #need to manually include fieldnames
    cw.writeheader()
    
    for rown, row in enumerate(NLX_data):
        try:
            #cw.writerow(row)
            cw.writerow(dict((k, v.encode('utf-8') if isinstance(v, unicode) else v) for k, v in row.iteritems())) #http://stackoverflow.com/questions/5838605/python-dictwriter-writing-utf-8-encoded-csv-files#comment6706041_5838817
        except:
            print rown
            print sys.exc_info()[0]              
            print sys.exc_info()[1]

